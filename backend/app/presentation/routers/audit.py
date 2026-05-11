"""
Router de auditoría y reporte histórico.

Endpoints:
    GET  /api/v1/audit          → lista paginada de despliegues
    GET  /api/v1/audit/export   → descarga Excel con el historial completo

Capa: Presentación
HU-15: Generación y descarga de reporte histórico de operación en Excel
"""

from __future__ import annotations

import io
import logging
from datetime import datetime, timezone

from fastapi import APIRouter, Query
from fastapi.responses import JSONResponse, StreamingResponse

from app.presentation.dependencies import audit_repository

logger = logging.getLogger(__name__)
router = APIRouter()


@router.get(
    "/audit",
    summary="Historial de despliegues",
    description="Retorna el listado paginado de todos los despliegues registrados.",
)
async def listar_historial(
    limite: int  = Query(default=50, ge=1, le=200),
    offset: int  = Query(default=0,  ge=0),
    estado: str | None = Query(
        default=None,
        description="Filtrar por estado: completed, failed, cancelled",
    ),
) -> JSONResponse:
    """Lista el historial de despliegues con paginación."""
    entradas = await audit_repository.listar(
        limite=limite, offset=offset, estado=estado
    )
    total = await audit_repository.contar(estado=estado)

    return JSONResponse({
        "total":   total,
        "limite":  limite,
        "offset":  offset,
        "entradas": [e.to_dict() for e in entradas],
    })


@router.get(
    "/audit/export",
    summary="Exportar historial como Excel",
    description=(
        "Genera y descarga un archivo .xlsx con el historial completo "
        "de despliegues. Útil para reportes institucionales y "
        "validación de tiempos operativos."
    ),
)
async def exportar_historial_excel() -> StreamingResponse:
    """
    Exporta el historial completo de despliegues como archivo Excel.

    El archivo incluye una hoja con el resumen y una hoja con
    el detalle de errores de los despliegues fallidos.
    """
    try:
        import openpyxl
        from openpyxl.styles import (
            Font, PatternFill, Alignment, Border, Side
        )
        from openpyxl.utils import get_column_letter
    except ImportError:
        return JSONResponse(
            status_code=500,
            content={"error": "openpyxl no está instalado."},
        )

    entradas = await audit_repository.listar(limite=10_000)

    wb = openpyxl.Workbook()

    # ── Hoja 1: Historial de despliegues ─────────────────────────────────
    ws = wb.active
    ws.title = "Historial de despliegues"

    # Estilos
    azul_poli    = "002D62"
    verde_ok     = "D4EDDA"
    rojo_error   = "F8D7DA"
    naranja_canc = "FFF3CD"
    gris_header  = "F8F9FA"

    estilo_header = Font(bold=True, color="FFFFFF", size=10)
    fill_header   = PatternFill("solid", fgColor=azul_poli)
    alineado_centro = Alignment(horizontal="center", vertical="center")
    borde_fino = Border(
        left=Side(style="thin",   color="DEE2E6"),
        right=Side(style="thin",  color="DEE2E6"),
        top=Side(style="thin",    color="DEE2E6"),
        bottom=Side(style="thin", color="DEE2E6"),
    )

    # Encabezados
    columnas = [
        ("Fecha",               16),
        ("Hora",                10),
        ("Curso",               38),
        ("ID Curso",            10),
        ("Plantilla ID",        12),
        ("Modelo",              14),
        ("Nivel",               12),
        ("Archivos",            10),
        ("Duración",            10),
        ("Estado",              16),
        ("ZIP",                 30),
        ("Task ID",             36),
    ]

    for col_idx, (titulo, ancho) in enumerate(columnas, start=1):
        celda = ws.cell(row=1, column=col_idx, value=titulo)
        celda.font      = estilo_header
        celda.fill      = fill_header
        celda.alignment = alineado_centro
        celda.border    = borde_fino
        ws.column_dimensions[get_column_letter(col_idx)].width = ancho

    ws.row_dimensions[1].height = 20
    ws.freeze_panes = "A2"

    # Filas de datos
    for fila_idx, entrada in enumerate(entradas, start=2):
        fecha_local = entrada.iniciado_en.astimezone(
            tz=None  # timezone local del servidor
        )

        valores = [
            fecha_local.strftime("%Y-%m-%d"),
            fecha_local.strftime("%H:%M:%S"),
            entrada.course_name,
            entrada.course_id,
            entrada.template_id,
            entrada.modelo_instruccional,
            entrada.nivel_formacion,
            entrada.archivos_subidos,
            entrada.duracion_display,
            entrada.estado_display,
            entrada.zip_filename,
            entrada.task_id,
        ]

        # Color de fila según estado
        color_fila = {
            "completed": verde_ok,
            "failed":    rojo_error,
            "cancelled": naranja_canc,
        }.get(entrada.estado, gris_header)

        fill_fila = PatternFill("solid", fgColor=color_fila)

        for col_idx, valor in enumerate(valores, start=1):
            celda = ws.cell(row=fila_idx, column=col_idx, value=valor)
            celda.fill   = fill_fila
            celda.border = borde_fino
            celda.alignment = Alignment(vertical="center")
            if col_idx in (1, 2, 4, 5, 8):
                celda.alignment = Alignment(
                    horizontal="center", vertical="center"
                )

    # ── Hoja 2: Resumen estadístico ───────────────────────────────────────
    ws2 = wb.create_sheet(title="Resumen")

    total    = len(entradas)
    exitosos = sum(1 for e in entradas if e.estado == "completed")
    fallidos = sum(1 for e in entradas if e.estado == "failed")
    cancelados = sum(1 for e in entradas if e.estado == "cancelled")
    duracion_promedio = (
        sum(e.duracion_seg for e in entradas if e.estado == "completed") / exitosos
        if exitosos > 0 else 0
    )
    total_archivos = sum(e.archivos_subidos for e in entradas)

    generado_en = datetime.now(timezone.utc).strftime("%Y-%m-%d %H:%M UTC")

    estadisticas = [
        ("Métrica",                    "Valor"),
        ("Total de despliegues",        total),
        ("Despliegues exitosos",        exitosos),
        ("Despliegues fallidos",        fallidos),
        ("Despliegues cancelados",      cancelados),
        ("Tasa de éxito",              f"{round(exitosos/total*100,1)}%" if total else "—"),
        ("Duración promedio (exitosos)", f"{int(duracion_promedio//60)}m {int(duracion_promedio%60)}s" if exitosos else "—"),
        ("Total archivos subidos",      total_archivos),
        ("Reporte generado",            generado_en),
    ]

    for fila_idx, (metrica, valor) in enumerate(estadisticas, start=1):
        celda_m = ws2.cell(row=fila_idx, column=1, value=metrica)
        celda_v = ws2.cell(row=fila_idx, column=2, value=valor)
        if fila_idx == 1:
            celda_m.font = Font(bold=True, color="FFFFFF")
            celda_v.font = Font(bold=True, color="FFFFFF")
            celda_m.fill = PatternFill("solid", fgColor=azul_poli)
            celda_v.fill = PatternFill("solid", fgColor=azul_poli)
        celda_m.border = borde_fino
        celda_v.border = borde_fino

    ws2.column_dimensions["A"].width = 32
    ws2.column_dimensions["B"].width = 20

    # ── Serializar y retornar ─────────────────────────────────────────────
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    nombre_archivo = (
        f"reporte_despliegues_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
    )

    logger.info(
        "Exportando historial Excel: %d registros → %s",
        total, nombre_archivo,
    )

    return StreamingResponse(
        buffer,
        media_type=(
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        ),
        headers={
            "Content-Disposition": f'attachment; filename="{nombre_archivo}"'
        },
    )