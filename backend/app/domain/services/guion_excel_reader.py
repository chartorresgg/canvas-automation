"""
Lector del archivo Excel de Guion de módulo.

Extrae las URLs de recursos multimedia y los párrafos de texto
que se inyectan en las páginas del aula virtual Canvas.

Estructura del Excel (columnas 0-3):
    col[0]: Etiqueta de sección  ("URL video inicial", "Unidad 1", "Cierre"...)
    col[1]: Etiqueta de subsección ("Inicio", "Video introducción...", "Párrafo de cierre"...)
    col[2]: Nombre/tipo de contenido (a menudo "No diligencie...\nU1_Lectura...")
    col[3]: Valor real (URL directa o texto, según la fila)

Capa: Dominio — services
Colaboradores: FrontPageComposer, MaterialFundamentalComposer
               (a través del orquestador)
"""

from __future__ import annotations

import logging
import re
from dataclasses import dataclass, field
from pathlib import Path

from openpyxl import load_workbook

logger = logging.getLogger(__name__)

_NOMBRE_HOJA = "Guion de módulo"

# Prefijos de instrucción que deben eliminarse del texto extraído
_PREFIJOS_INSTRUCCION = [
    "no diligencie nada en este espacio",
    "solo diligencie el nombre del módulo",
    "solo diligencie el nombre del modulo",
]


# ──────────────────────────────────────────────────────────────────────────────
# Value Objects de resultado
# ──────────────────────────────────────────────────────────────────────────────

@dataclass
class UnidadGuion:
    """Datos extraídos del Guion Excel para una unidad del módulo."""
    numero:             int
    video_intro_url:    str | None = None
    parrafo_intro:      str        = ""
    podcast_url:        str | None = None
    vimeo_mat_fund_url: str | None = None


@dataclass
class GuionData:
    """
    Datos completos extraídos del Guion Excel de un módulo.

    Agrupa todos los recursos multimedia y textos que deben
    inyectarse en el front del curso y las páginas de Material
    Fundamental.
    """
    video_inicial_url: str | None             = None
    texto_inicio:      str                    = ""
    texto_cierre:      str                    = ""
    unidades:          dict[int, UnidadGuion] = field(default_factory=dict)

    def unidad(self, numero: int) -> UnidadGuion:
        """Retorna los datos de una unidad o un objeto vacío si no existe."""
        return self.unidades.get(numero, UnidadGuion(numero=numero))


# ──────────────────────────────────────────────────────────────────────────────
# Lector principal
# ──────────────────────────────────────────────────────────────────────────────

class GuionExcelReader:
    """
    Lee y parsea el archivo Excel 'Guion de módulo'.

    Responsabilidad única (SRP): extraer datos del Excel de contenidos
    y exponerlos como GuionData para que los Composers los usen.

    Estrategia de parsing por posición de columna:
        col[0] → etiqueta de sección principal
        col[1] → etiqueta de subsección
        col[2] → nombre/tipo de contenido (a menudo instrucción + valor)
        col[3] → valor real (URL directa, texto largo o URL embed)

    Colaboradores: ninguno (servicio de dominio puro).
    """

    def __init__(self, excel_path: Path) -> None:
        self._path = excel_path

    def read(self) -> GuionData:
        """
        Lee el Excel y retorna el GuionData completo.

        Returns:
            GuionData con todos los recursos multimedia y textos.

        Raises:
            FileNotFoundError: Si el archivo Excel no existe.
            ValueError:        Si la hoja esperada no se encuentra.
        """
        if not self._path.exists():
            raise FileNotFoundError(
                f"Guion Excel no encontrado: '{self._path}'"
            )

        wb = load_workbook(str(self._path), read_only=True, data_only=True)

        ws = None
        for nombre in wb.sheetnames:
            if _NOMBRE_HOJA.lower() in nombre.lower():
                ws = wb[nombre]
                break
        if ws is None:
            ws = wb.active
            logger.warning(
                "No se encontró hoja '%s'. Usando hoja activa: '%s'",
                _NOMBRE_HOJA, ws.title,
            )

        filas = self._leer_filas(ws)
        wb.close()

        datos = self._parsear(filas)

        logger.info(
            "Guion leído — video_inicial=%s | unidades=%s | cierre=%s",
            bool(datos.video_inicial_url),
            {u: {
                "video_intro": bool(v.video_intro_url),
                "podcast":     bool(v.podcast_url),
                "vimeo":       bool(v.vimeo_mat_fund_url),
                "parrafo":     bool(v.parrafo_intro),
            } for u, v in datos.unidades.items()},
            bool(datos.texto_cierre),
        )
        return datos

    # ──────────────────────────────────────────────────────────────
    # Privados — lectura
    # ──────────────────────────────────────────────────────────────

    @staticmethod
    def _leer_filas(ws) -> list[list[str]]:
        """
        Convierte el worksheet en lista de filas con strings limpios.
        Ignora filas completamente vacías.
        """
        filas: list[list[str]] = []
        for fila in ws.iter_rows(values_only=True):
            celdas = [str(c).strip() if c is not None else "" for c in fila]
            if any(c for c in celdas):
                filas.append(celdas)
        return filas

    def _parsear(self, filas: list[list[str]]) -> GuionData:
        """
        Parsea las filas del Excel usando la posición de columna como guía.

        Estrategia por sección:
            · URL video inicial → col[0] == "URL video inicial" → URL en col[1]
            · Texto inicio      → col[1] == "Inicio"            → texto en col[2]
            · Unidad activa     → col[0] == "Unidad N"
            · Video intro       → col[1] == "Video introducción..."  → URL en col[3]
            · Párrafo intro     → col[1] == "Párrafo de introducción" → texto en col[2]
            · Multimedia MF     → col[2] contiene "material_fundamental_1"
                                   → URL en col[3] de esta fila O de la siguiente
            · Párrafo cierre    → col[1] == "Párrafo de cierre"  → texto en col[2]
        """
        datos = GuionData()
        unidad_actual: int | None = None

        for idx, fila in enumerate(filas):
            # Garantizar mínimo 4 columnas
            while len(fila) < 4:
                fila.append("")

            col0 = fila[0].strip()
            col1 = fila[1].strip()
            col2 = fila[2].strip()
            col3 = fila[3].strip()

            col0_l = col0.lower()
            col1_l = col1.lower()
            col2_l = col2.lower().replace(" ", "_")

            # ── 1. URL del video inicial ──────────────────────────
            if col0_l == "url video inicial":
                url = self._extraer_url(col1)
                if url:
                    datos.video_inicial_url = url
                    logger.info("Video inicial: %s", url)

            # ── 2. Texto de inicio ────────────────────────────────
            if col1_l == "inicio" and not datos.texto_inicio:
                texto = self._limpiar_texto(col2)
                if texto:
                    datos.texto_inicio = texto
                    logger.info("Texto inicio: %s chars", len(texto))

            for u in range(1, 5):
                if f"unidad {u}" in col0_l:
                    unidad_actual = u
                    if u not in datos.unidades:
                        datos.unidades[u] = UnidadGuion(numero=u)
                    break

            # ── 4. Datos por unidad (solo si hay unidad activa) ───
            if unidad_actual is not None:
                unidad = datos.unidades.setdefault(
                    unidad_actual, UnidadGuion(numero=unidad_actual)
                )

                # Video de introducción — URL está en col[3]
                if ("video introducción" in col1_l or
                        "video introduccion" in col1_l):
                    url = self._extraer_url(col3)
                    if url and unidad.video_intro_url is None:
                        unidad.video_intro_url = url
                        logger.info("U%d video_intro: %s", unidad_actual, url)

                # Párrafo de introducción — texto en col[2]
                if ("párrafo de introducción" in col1_l or
                        "parrafo de introduccion" in col1_l):
                    texto = self._limpiar_texto(col2)
                    if texto and not unidad.parrafo_intro:
                        unidad.parrafo_intro = texto
                        logger.info(
                            "U%d parrafo_intro: %s chars",
                            unidad_actual, len(texto),
                        )

                # Material fundamental multimedia
                # La etiqueta "material_fundamental_1" aparece en col[2]
                # La URL puede estar en col[3] de esta fila o de la siguiente
                if "material_fundamental_1" in col2_l:
                    url = self._extraer_url(col3)
                    if url:
                        self._asignar_multimedia(url, unidad)
                    elif idx + 1 < len(filas):
                        sig = filas[idx + 1]
                        url_sig = self._extraer_url(
                            sig[3] if len(sig) > 3 else ""
                        )
                        if url_sig:
                            self._asignar_multimedia(url_sig, unidad)

            # ── 5. Párrafo de cierre ──────────────────────────────
            if ("párrafo de cierre" in col1_l or
                    "parrafo de cierre" in col1_l):
                texto = self._limpiar_texto(col2)
                if texto and not datos.texto_cierre:
                    datos.texto_cierre = texto
                    logger.info("Texto cierre: %s chars", len(texto))

        return datos

    # ──────────────────────────────────────────────────────────────
    # Privados — utilidades
    # ──────────────────────────────────────────────────────────────

    @staticmethod
    def _extraer_url(texto: str) -> str | None:
        """
        Extrae la primera URL HTTP/HTTPS de un texto.
        Elimina URLs de Shutterstock (son imágenes de referencia, no recursos).

        Returns:
            URL limpia o None si no se encuentra o es de Shutterstock.
        """
        if not texto:
            return None
        match = re.search(r"https?://\S+", texto)
        if not match:
            return None
        url = match.group(0).rstrip(".,;)")
        if "shutterstock" in url.lower():
            return None
        return url

    @staticmethod
    def _limpiar_texto(texto: str, min_chars: int = 30) -> str:
        """
        Limpia un texto de celda extrayendo solo el contenido real.

        Elimina prefijos de instrucción institucionales como:
            "No diligencie nada en este espacio"
            "Solo diligencie el nombre del módulo"
        que siempre preceden al contenido real separados por \\n.

        Args:
            texto:     Contenido de la celda.
            min_chars: Longitud mínima del resultado para ser válido.

        Returns:
            Texto limpio o cadena vacía si no hay contenido válido.
        """
        if not texto:
            return ""

        # Separar por saltos de línea y filtrar líneas de instrucción
        lineas = texto.split("\n")
        lineas_validas = []

        for linea in lineas:
            linea = linea.strip()
            if not linea:
                continue
            linea_lower = linea.lower()
            es_instruccion = any(
                pref in linea_lower for pref in _PREFIJOS_INSTRUCCION
            )
            if not es_instruccion:
                lineas_validas.append(linea)

        resultado = " ".join(lineas_validas).strip()
        return resultado if len(resultado) >= min_chars else ""

    @staticmethod
    def _asignar_multimedia(url: str, unidad: UnidadGuion) -> None:
        """
        Clasifica una URL como podcast (SoundCloud) o video (Vimeo)
        y la asigna al campo correspondiente de la unidad.

        Ignora URLs vacías o de N/A.
        """
        if not url or url.strip().upper() in ("N/A", "NA", ""):
            return

        url_lower = url.lower()
        if "soundcloud" in url_lower:
            unidad.podcast_url = url
            logger.info("U%d podcast: %s", unidad.numero, url[:70])
        elif "vimeo" in url_lower or "player" in url_lower:
            unidad.vimeo_mat_fund_url = url
            logger.info("U%d vimeo_mat_fund: %s", unidad.numero, url[:70])